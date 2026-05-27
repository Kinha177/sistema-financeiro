from __future__ import annotations
from decimal import Decimal
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter


# ── constantes de estilo ──────────────────────────────────────────────────────

_FILL_HEADER = PatternFill("solid", fgColor="1a2e4a")
_FILL_SECAO  = PatternFill("solid", fgColor="2d4a6e")
_FILL_TOTAL  = PatternFill("solid", fgColor="dce6f0")
_FILL_ALT    = PatternFill("solid", fgColor="f5f7fa")
_FILL_LUCRO  = PatternFill("solid", fgColor="d1fae5")
_FILL_PREJUIZO = PatternFill("solid", fgColor="fee2e2")

_FONT_HEADER = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_FONT_BOLD   = Font(bold=True, name="Calibri", size=10)
_FONT_NORMAL = Font(name="Calibri", size=10)
_FONT_TOTAL  = Font(bold=True, name="Calibri", size=10, color="1a2e4a")
_FONT_TITLE  = Font(bold=True, name="Calibri", size=14, color="1a2e4a")
_FONT_SUBTITLE = Font(name="Calibri", size=10, color="6b7280")

_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

_THIN = Side(style="thin", color="d1d5db")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ── helpers globais ───────────────────────────────────────────────────────────

def _brl(v) -> str:
    if v is None:
        return "R$ 0,00"
    d = Decimal(str(v))
    neg = d < 0
    d = abs(d)
    inteiro, *frac = f"{d:.2f}".split(".")
    dec = frac[0] if frac else "00"
    grupos: list[str] = []
    s = inteiro
    while len(s) > 3:
        grupos.insert(0, s[-3:])
        s = s[:-3]
    grupos.insert(0, s)
    formatted = ".".join(grupos) + "," + dec
    return f"R$ -{formatted}" if neg else f"R$ {formatted}"


def _dt(v: date | None) -> str:
    return v.strftime("%d/%m/%Y") if v else ""


class ExcelService:
    """Geração de planilhas Excel com formatação profissional."""

    def __init__(self, empresa_nome: str = "Empresa") -> None:
        self.empresa_nome = empresa_nome

    # ── helpers de célula ─────────────────────────────────────────────────────

    def _cell(self, ws, row: int, col: int, value,
              font=None, fill=None, align=None, border=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font   = font  or _FONT_NORMAL
        c.fill   = fill  or PatternFill()
        c.alignment = align or _ALIGN_LEFT
        if border:
            c.border = _BORDER
        return c

    def _header_row(self, ws, row: int, labels: list[str]) -> None:
        for col, label in enumerate(labels, 1):
            self._cell(ws, row, col, label, font=_FONT_HEADER, fill=_FILL_HEADER,
                       align=_ALIGN_CENTER)

    def _title_block(self, ws, titulo: str, subtitulo: str = "") -> int:
        """Insere título + subtítulo e retorna a próxima linha disponível."""
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        c = ws.cell(row=1, column=1, value=f"{self.empresa_nome}  —  {titulo}")
        c.font      = _FONT_TITLE
        c.alignment = _ALIGN_LEFT
        if subtitulo:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
            c2 = ws.cell(row=2, column=1, value=subtitulo)
            c2.font      = _FONT_SUBTITLE
            c2.alignment = _ALIGN_LEFT
            ws.row_dimensions[2].height = 16
            return 4
        return 3

    def _auto_width(self, ws, min_w: int = 10, max_w: int = 60) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value else ""
                    max_len = max(max_len, len(val))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))

    def _alt_fill(self, row_idx: int) -> PatternFill | None:
        return _FILL_ALT if row_idx % 2 == 0 else None

    # ══════════════════════════════════════════════════════════════════════════
    # Livro Diário
    # ══════════════════════════════════════════════════════════════════════════

    def exportar_livro_diario(
        self,
        dados: list[dict],
        data_inicio: date,
        data_fim: date,
        destino: Path,
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Livro Diário"

        subtitulo = f"{_dt(data_inicio)} a {_dt(data_fim)}"
        first_data_row = self._title_block(ws, "Livro Diário", subtitulo)

        cols = ["Data", "Nº Lançamento", "Histórico", "Conta Débito", "Conta Crédito", "Valor (R$)"]
        self._header_row(ws, first_data_row, cols)

        r = first_data_row + 1
        total = Decimal("0")
        for lanc in dados:
            data_str = _dt(lanc.get("data"))
            historico = lanc.get("historico", "")
            lanc_id   = lanc.get("id", "")
            for idx, item in enumerate(lanc.get("itens", [])):
                valor = Decimal(str(item.get("valor", 0)))
                total += valor
                fill = self._alt_fill(r)
                self._cell(ws, r, 1, data_str if idx == 0 else "",         fill=fill, align=_ALIGN_CENTER)
                self._cell(ws, r, 2, str(lanc_id) if idx == 0 else "",     fill=fill, align=_ALIGN_CENTER)
                self._cell(ws, r, 3, historico if idx == 0 else "",        fill=fill, align=_ALIGN_LEFT)
                self._cell(ws, r, 4, item.get("conta_debito",  ""),        fill=fill)
                self._cell(ws, r, 5, item.get("conta_credito", ""),        fill=fill)
                self._cell(ws, r, 6, _brl(valor),                          fill=fill, align=_ALIGN_RIGHT)
                r += 1

        # rodapé total
        for col in range(1, 6):
            self._cell(ws, r, col, "", font=_FONT_BOLD, fill=_FILL_TOTAL)
        self._cell(ws, r, 5, "TOTAL", font=_FONT_TOTAL, fill=_FILL_TOTAL, align=_ALIGN_RIGHT)
        self._cell(ws, r, 6, _brl(total), font=_FONT_TOTAL, fill=_FILL_TOTAL, align=_ALIGN_RIGHT)

        self._auto_width(ws)
        ws.freeze_panes = f"A{first_data_row + 1}"
        wb.save(str(destino))
        return destino

    # ══════════════════════════════════════════════════════════════════════════
    # Livro Razão
    # ══════════════════════════════════════════════════════════════════════════

    def exportar_livro_razao(
        self,
        dados: list[dict],
        conta_nome: str,
        destino: Path,
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Livro Razão"

        first_data_row = self._title_block(ws, "Livro Razão", conta_nome)

        cols = ["Data", "Histórico", "Débito (R$)", "Crédito (R$)", "Saldo (R$)"]
        self._header_row(ws, first_data_row, cols)

        r = first_data_row + 1
        for mov in dados:
            fill = self._alt_fill(r)
            self._cell(ws, r, 1, _dt(mov.get("data")),         fill=fill, align=_ALIGN_CENTER)
            self._cell(ws, r, 2, mov.get("historico", ""),     fill=fill)
            self._cell(ws, r, 3, _brl(mov.get("debito")),      fill=fill, align=_ALIGN_RIGHT)
            self._cell(ws, r, 4, _brl(mov.get("credito")),     fill=fill, align=_ALIGN_RIGHT)
            self._cell(ws, r, 5, _brl(mov.get("saldo")),       fill=fill, align=_ALIGN_RIGHT)
            r += 1

        if dados:
            saldo_final = dados[-1].get("saldo", Decimal("0"))
            for col in [1, 2, 3, 4]:
                self._cell(ws, r, col, "", font=_FONT_BOLD, fill=_FILL_TOTAL)
            self._cell(ws, r, 2, "SALDO FINAL", font=_FONT_TOTAL, fill=_FILL_TOTAL)
            self._cell(ws, r, 5, _brl(saldo_final), font=_FONT_TOTAL, fill=_FILL_TOTAL, align=_ALIGN_RIGHT)

        self._auto_width(ws)
        ws.freeze_panes = f"A{first_data_row + 1}"
        wb.save(str(destino))
        return destino

    # ══════════════════════════════════════════════════════════════════════════
    # DRE
    # ══════════════════════════════════════════════════════════════════════════

    def exportar_dre(
        self,
        dados: dict,
        data_inicio: date,
        data_fim: date,
        destino: Path,
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "DRE"

        subtitulo = f"{_dt(data_inicio)} a {_dt(data_fim)}"
        r = self._title_block(ws, "DRE — Demonstração do Resultado do Exercício", subtitulo)

        def _secao(titulo: str, itens: list[dict], total: Decimal, total_label: str):
            nonlocal r
            # cabeçalho da seção
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            c = ws.cell(row=r, column=1, value=titulo)
            c.font  = _FONT_HEADER
            c.fill  = _FILL_SECAO
            c.alignment = _ALIGN_LEFT
            r += 1

            self._header_row(ws, r, ["Código", "Descrição", "Valor (R$)"])
            r += 1

            for idx, item in enumerate(itens):
                fill = self._alt_fill(idx)
                self._cell(ws, r, 1, item.get("codigo", ""), fill=fill, align=_ALIGN_CENTER)
                self._cell(ws, r, 2, item.get("nome",   ""), fill=fill)
                self._cell(ws, r, 3, _brl(item.get("valor")), fill=fill, align=_ALIGN_RIGHT)
                r += 1

            self._cell(ws, r, 1, "", font=_FONT_BOLD, fill=_FILL_TOTAL)
            self._cell(ws, r, 2, total_label, font=_FONT_TOTAL, fill=_FILL_TOTAL)
            self._cell(ws, r, 3, _brl(total), font=_FONT_TOTAL, fill=_FILL_TOTAL, align=_ALIGN_RIGHT)
            r += 2

        _secao("RECEITAS OPERACIONAIS",
               dados.get("receitas", []),
               dados.get("total_receitas", Decimal("0")),
               "Total de Receitas")
        _secao("CUSTOS",
               dados.get("custos", []),
               dados.get("total_custos", Decimal("0")),
               "Total de Custos")
        _secao("DESPESAS OPERACIONAIS",
               dados.get("despesas", []),
               dados.get("total_despesas", Decimal("0")),
               "Total de Despesas")

        # resultado
        lucro = dados.get("lucro_liquido", Decimal("0"))
        label = "LUCRO LÍQUIDO" if lucro >= 0 else "PREJUÍZO LÍQUIDO"
        fill_result = _FILL_LUCRO if lucro >= 0 else _FILL_PREJUIZO
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        self._cell(ws, r, 1, label,      font=Font(bold=True, size=11, name="Calibri"), fill=fill_result)
        self._cell(ws, r, 3, _brl(lucro), font=Font(bold=True, size=11, name="Calibri"), fill=fill_result, align=_ALIGN_RIGHT)

        self._auto_width(ws)
        wb.save(str(destino))
        return destino

    # ══════════════════════════════════════════════════════════════════════════
    # Balanço Patrimonial
    # ══════════════════════════════════════════════════════════════════════════

    def exportar_balanco(
        self,
        dados: dict,
        data_base: date,
        destino: Path,
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Balanço Patrimonial"

        r = self._title_block(ws, "Balanço Patrimonial", _dt(data_base))

        # Layout: A=cód_ativo, B=nome_ativo, C=val_ativo, D=espaço, E=cód_pass, F=nome_pass, G=val_pass
        header_labels = ["Código", "ATIVO", "Valor (R$)", "", "Código", "PASSIVO + PL", "Valor (R$)"]
        for col, label in enumerate(header_labels, 1):
            if col == 4:
                ws.cell(row=r, column=4, value="")
                continue
            c = ws.cell(row=r, column=col, value=label)
            c.font  = _FONT_HEADER
            c.fill  = _FILL_HEADER if col != 4 else PatternFill()
            c.alignment = _ALIGN_CENTER
            c.border = _BORDER
        r += 1

        ativo   = dados.get("ativo",   [])
        passivo = dados.get("passivo", [])
        pl      = dados.get("pl",      [])
        passivo_pl = passivo + [{"codigo": "—", "nome": "— Patrimônio Líquido —", "valor": None}] + pl

        max_rows = max(len(ativo), len(passivo_pl))

        for i in range(max_rows):
            fill_a = self._alt_fill(i)
            fill_p = self._alt_fill(i)

            # ativo
            if i < len(ativo):
                item_a = ativo[i]
                self._cell(ws, r, 1, item_a.get("codigo", ""), fill=fill_a, align=_ALIGN_CENTER)
                self._cell(ws, r, 2, item_a.get("nome",   ""), fill=fill_a)
                self._cell(ws, r, 3, _brl(item_a.get("valor")), fill=fill_a, align=_ALIGN_RIGHT)
            else:
                for col in [1, 2, 3]:
                    self._cell(ws, r, col, "", fill=fill_a)

            ws.cell(row=r, column=4, value="")  # spacer

            # passivo + pl
            if i < len(passivo_pl):
                item_p = passivo_pl[i]
                val = item_p.get("valor")
                is_section = val is None
                fnt = _FONT_BOLD if is_section else _FONT_NORMAL
                self._cell(ws, r, 5, item_p.get("codigo", ""), font=fnt, fill=fill_p, align=_ALIGN_CENTER)
                self._cell(ws, r, 6, item_p.get("nome",   ""), font=fnt, fill=fill_p)
                self._cell(ws, r, 7, _brl(val) if not is_section else "", font=fnt, fill=fill_p, align=_ALIGN_RIGHT)
            else:
                for col in [5, 6, 7]:
                    self._cell(ws, r, col, "", fill=fill_p)
            r += 1

        # totais
        for col in [1, 2]:
            self._cell(ws, r, col, "", font=_FONT_BOLD, fill=_FILL_TOTAL)
        self._cell(ws, r, 2, "TOTAL ATIVO",       font=_FONT_TOTAL, fill=_FILL_TOTAL)
        self._cell(ws, r, 3, _brl(dados.get("total_ativo")), font=_FONT_TOTAL, fill=_FILL_TOTAL, align=_ALIGN_RIGHT)
        ws.cell(row=r, column=4, value="")
        for col in [5, 6]:
            self._cell(ws, r, col, "", font=_FONT_BOLD, fill=_FILL_TOTAL)
        self._cell(ws, r, 6, "TOTAL PASSIVO + PL", font=_FONT_TOTAL, fill=_FILL_TOTAL)
        self._cell(ws, r, 7, _brl(dados.get("total_passivo_pl")), font=_FONT_TOTAL, fill=_FILL_TOTAL, align=_ALIGN_RIGHT)

        if not dados.get("equacao_valida", True):
            r += 2
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            c = ws.cell(row=r, column=1,
                        value="⚠ Equação patrimonial desequilibrada — verifique os lançamentos.")
            c.font = Font(bold=True, color="DC2626", name="Calibri", size=10)

        self._auto_width(ws)
        wb.save(str(destino))
        return destino

    # ══════════════════════════════════════════════════════════════════════════
    # Razonetes
    # ══════════════════════════════════════════════════════════════════════════

    def exportar_razonetes(self, dados: list[dict], destino: Path) -> Path:
        """
        dados: lista de dicts com conta_nome, lado_debito, lado_credito, saldo_final.
        Organiza 2 razonetes por linha, cada um em 2 colunas (Débito | Crédito).
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Razonetes"

        r = self._title_block(ws, "Razonetes", "Contas em Forma de T")

        # Cada razonete ocupa 2 colunas + 1 espaçadora.
        # Par de razonetes: colunas 1-2 | 4-5  (3 e 6 são espaçadoras)
        def _write_razonete(dados_rz: dict, col_start: int, row_start: int) -> int:
            """Escreve um razonete e retorna a quantidade de linhas ocupadas."""
            conta_nome = dados_rz.get("conta_nome", "—")
            lado_deb   = dados_rz.get("lado_debito",  [])
            lado_cred  = dados_rz.get("lado_credito", [])
            saldo      = dados_rz.get("saldo_final",  Decimal("0"))

            # cabeçalho (nome da conta)
            ws.merge_cells(
                start_row=row_start, start_column=col_start,
                end_row=row_start,   end_column=col_start + 1,
            )
            c = ws.cell(row=row_start, column=col_start, value=conta_nome)
            c.font      = _FONT_HEADER
            c.fill      = _FILL_HEADER
            c.alignment = _ALIGN_CENTER
            c.border    = _BORDER
            ws.cell(row=row_start, column=col_start + 1).border = _BORDER

            row = row_start + 1

            # sub-cabeçalho Débito / Crédito
            for ci, label in enumerate(["Débito (R$)", "Crédito (R$)"]):
                c = ws.cell(row=row, column=col_start + ci, value=label)
                c.font      = _FONT_HEADER
                c.fill      = PatternFill("solid", fgColor="2d4a6e")
                c.alignment = _ALIGN_CENTER
                c.border    = _BORDER
            row += 1

            max_rows = max(len(lado_deb), len(lado_cred), 1)
            for i in range(max_rows):
                fill = self._alt_fill(i)
                deb_val  = _brl(lado_deb[i]["valor"])  if i < len(lado_deb)  else ""
                cred_val = _brl(lado_cred[i]["valor"]) if i < len(lado_cred) else ""
                self._cell(ws, row, col_start,     deb_val,  fill=fill, align=_ALIGN_RIGHT)
                self._cell(ws, row, col_start + 1, cred_val, fill=fill, align=_ALIGN_RIGHT)
                row += 1

            # linha de saldo
            self._cell(ws, row, col_start,     "Saldo",    font=_FONT_BOLD,  fill=_FILL_TOTAL)
            self._cell(ws, row, col_start + 1, _brl(saldo), font=_FONT_TOTAL, fill=_FILL_TOTAL, align=_ALIGN_RIGHT)
            row += 1

            return row - row_start  # linhas consumidas

        pairs = [dados[i:i + 2] for i in range(0, len(dados), 2)]
        for pair in pairs:
            lines_used = 0
            for idx, rz in enumerate(pair):
                col_start = 1 if idx == 0 else 4
                used = _write_razonete(rz, col_start, r)
                lines_used = max(lines_used, used)
            r += lines_used + 1  # espaço entre pares

        self._auto_width(ws)
        wb.save(str(destino))
        return destino

    # ══════════════════════════════════════════════════════════════════════════
    # Estoque / PEPS-UEPS
    # ══════════════════════════════════════════════════════════════════════════

    def exportar_ficha_peps_ueps(
        self,
        dados: list[dict],
        produto_nome: str,
        metodo: str,
        destino: Path,
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Custeio {metodo}"

        r = self._title_block(ws, f"Custeio {metodo}", produto_nome)

        self._header_row(ws, r, ["Lote", "Data", "Quantidade", "Valor Unit. (R$)", "Valor Total (R$)"])
        r += 1

        grand_total = Decimal("0")
        for i, lote in enumerate(dados, 1):
            fill = self._alt_fill(i)
            vt = Decimal(str(lote.get("valor_total", 0)))
            grand_total += vt
            self._cell(ws, r, 1, str(i),                        fill=fill, align=_ALIGN_CENTER)
            self._cell(ws, r, 2, _dt(lote.get("data")),         fill=fill, align=_ALIGN_CENTER)
            self._cell(ws, r, 3, str(lote.get("quantidade", "")), fill=fill, align=_ALIGN_RIGHT)
            self._cell(ws, r, 4, _brl(lote.get("valor_unitario")), fill=fill, align=_ALIGN_RIGHT)
            self._cell(ws, r, 5, _brl(vt),                      fill=fill, align=_ALIGN_RIGHT)
            r += 1

        for col in [1, 2, 3, 4]:
            self._cell(ws, r, col, "", font=_FONT_BOLD, fill=_FILL_TOTAL)
        self._cell(ws, r, 4, "TOTAL",           font=_FONT_TOTAL, fill=_FILL_TOTAL)
        self._cell(ws, r, 5, _brl(grand_total), font=_FONT_TOTAL, fill=_FILL_TOTAL, align=_ALIGN_RIGHT)

        self._auto_width(ws)
        ws.freeze_panes = f"A{r - len(dados)}"
        wb.save(str(destino))
        return destino
